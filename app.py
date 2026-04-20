import os
import math
import uuid
import time
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
import psycopg2
from psycopg2 import pool
from functools import wraps

print(f"[STARTUP] PORT env = {os.getenv('PORT')}")
print(f"[STARTUP] DATABASE_URL set = {bool(os.getenv('DATABASE_URL'))}")

# ================= APP =================
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev")
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# ================= FOLDER =================
PROFILE_UPLOAD_FOLDER = 'static/uploads'
SHAPE_UPLOAD_FOLDER   = 'uploads/shapefile'

os.makedirs(PROFILE_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(SHAPE_UPLOAD_FOLDER,   exist_ok=True)

app.config['PROFILE_UPLOAD_FOLDER'] = PROFILE_UPLOAD_FOLDER
app.config['SHAPE_UPLOAD_FOLDER']   = SHAPE_UPLOAD_FOLDER

# ================= CONNECTION POOL =================
# Buka koneksi sekali saat startup, reuse untuk semua request
# connect_timeout=5 → tidak hanging 30+ detik kalau DB lambat
DATABASE_URL = os.getenv("DATABASE_URL")

# fallback untuk lokal (opsional tapi disarankan)
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL environment variable is not set!")

db_pool = None

def get_db_pool():
    global db_pool
    if db_pool is None:
        db_pool = psycopg2.pool.SimpleConnectionPool(
            minconn=1,
            maxconn=10,
            dsn=DATABASE_URL,
            connect_timeout=5
        )
    return db_pool

def get_db_connection():
    return get_db_pool().getconn()

def release_db_connection(conn):
    get_db_pool().putconn(conn)

# ================= VALIDASI =================
ALLOWED_IMAGE_EXT = {'png', 'jpg', 'jpeg'}
ALLOWED_FILE_EXT  = {'xlsx', 'xls'}

def allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXT

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_FILE_EXT

# ================= ROLE =================
def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def wrap(*args, **kwargs):
            if 'role' not in session:
                return redirect(url_for('login'))
            if session['role'] not in allowed_roles:
                return render_template('403.html'), 403
            return f(*args, **kwargs)
        return wrap
    return decorator

# ================= AUTH =================
@app.route('/')
@app.route('/welcome')
def welcome():
    return render_template('welcome.html')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'}), 200

# ================= LOGIN =================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email    = request.form.get('email')
        password = request.form.get('password')

        # validasi input
        if not email or not password:
            flash("Email dan password wajib diisi!")
            return redirect(url_for('login'))

        conn = get_db_connection()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, role, full_name, password, profile_image
                FROM users WHERE email = %s
            """, (email,))
            user = cur.fetchone()
            cur.close()
        except Exception as e:
            print("Login error:", e)
            flash("Terjadi kesalahan pada server!")
            return redirect(url_for('login'))
        finally:
            release_db_connection(conn)

        # cek password
        if user and check_password_hash(user[3], password):
            session['user_id']       = user[0]
            session['role']          = user[1]
            session['full_name']     = user[2]
            session['email']         = email
            session['profile_image'] = user[4]

            return redirect(url_for('main'))
        else:
            flash("Email atau password salah!")
            return redirect(url_for('login'))

    return render_template('login.html')


# ================= SIGNUP =================
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email     = request.form.get('email')
        password  = request.form.get('password')
        role      = request.form.get('role')

        # validasi input
        if not full_name or not email or not password or not role:
            flash("Semua field wajib diisi!")
            return redirect(url_for('signup'))

        conn = get_db_connection()
        try:
            cur = conn.cursor()

            # cek email sudah ada atau belum
            cur.execute("SELECT id FROM users WHERE email = %s", (email,))
            if cur.fetchone():
                flash("Email sudah terdaftar!")
                return redirect(url_for('signup'))

            # hash password
            hashed_password = generate_password_hash(password)

            # insert user
            cur.execute("""
                INSERT INTO users (full_name, email, password, role)
                VALUES (%s, %s, %s, %s)
            """, (full_name, email, hashed_password, role))

            conn.commit()
            cur.close()

        except Exception as e:
            print("Signup error:", e)
            flash("Terjadi kesalahan saat registrasi!")
            return redirect(url_for('signup'))
        finally:
            release_db_connection(conn)

        flash("Registrasi berhasil! Silakan login.")
        return redirect(url_for('login'))

    return render_template('signup.html')

# ================= PROFILE =================
@app.route('/profile')
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT full_name, email, role FROM users WHERE id = %s", (session['user_id'],))
        user = cur.fetchone()
        cur.close()
    finally:
        release_db_connection(conn)

    return render_template('profile.html', user=user)

@app.route('/upload_profile', methods=['POST'])
def upload_profile():
    file = request.files['profile_image']

    if file and allowed_image(file.filename):
        ext      = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{uuid.uuid4()}.{ext}"
        file.save(os.path.join(app.config['PROFILE_UPLOAD_FOLDER'], filename))

        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE users SET profile_image = %s WHERE id = %s",
                (filename, session['user_id'])
            )
            conn.commit()
            cursor.close()
        finally:
            release_db_connection(conn)

        session['profile_image'] = filename

    return redirect(url_for('profile'))

@app.route('/change-password', methods=['POST'])
def change_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    old_password     = request.form['old_password']
    new_password     = request.form['new_password']
    confirm_password = request.form['confirm_password']

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT password FROM users WHERE id = %s", (session['user_id'],))
        user = cur.fetchone()

        # 1. Cek Password Lama
        if not user or not check_password_hash(user[0], old_password):
            flash("Password lama salah!", "danger") # Cukup satu kali, beri kategori danger
            cur.close()
            return redirect(url_for('profile'))

        # 2. Cek Konfirmasi Password Baru
        if new_password != confirm_password:
            flash("Konfirmasi password tidak cocok!", "danger") # Cukup satu kali, beri kategori danger
            cur.close()
            return redirect(url_for('profile'))

        # 3. Update Database jika semua valid
        hashed_password = generate_password_hash(new_password)
        cur.execute(
            "UPDATE users SET password = %s WHERE id = %s",
            (hashed_password, session['user_id'])
        )
        conn.commit()
        cur.close()
        
        flash("Password berhasil diubah!", "success") # Kategori success untuk warna hijau

    finally:
        release_db_connection(conn)

    return redirect(url_for('profile'))

# ================= PAGE ROUTES =================
@app.route('/main')
@role_required(['dt_engineer', 'rf_engineer'])
def main():
    start = time.time()
    result = render_template('main.html')
    print(f"[TIMER] /main render: {time.time() - start:.3f}s")
    return result

@app.route('/route')
@role_required(['dt_engineer', 'rf_engineer'])
def route():
    return render_template('route.html')

@app.route('/drivetest')
@role_required(['dt_engineer', 'rf_engineer'])
def drivetest():
    return render_template('drivetest.html')

@app.route('/coverage')
@role_required(['rf_engineer'])
def coverage():
    return render_template('coverage.html')

@app.route('/analysis')
@role_required(['rf_engineer'])
def analysis():
    return render_template('analysis.html')

@app.route('/evaluation')
@role_required(['rf_engineer'])
def evaluation():
    return render_template('evaluation.html')

@app.route('/newsite')
@role_required(['rf_engineer'])
def newsite():
    return render_template('newsite.html')

@app.route('/simulation_dt')
@role_required(['rf_engineer'])
def simulation_dt():
    return render_template('simulation_dt.html')

@app.route('/help')
def help_page():
    return render_template('help.html')

@app.route('/about')
def about():
    return render_template('about.html')

# ================= API: UPLOAD SITE XLSX =================
@app.route('/api/upload-site', methods=['POST'])
def upload_site():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang dikirim'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Nama file kosong'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Format file harus .xlsx atau .xls'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['SHAPE_UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        site_index = parse_xlsx(filepath)
    except Exception as e:
        return jsonify({'error': f'Gagal parsing XLSX: {str(e)}'}), 500

    if not site_index:
        return jsonify({'error': 'Tidak ada data site valid di file ini'}), 400

    # Simpan nama file ke session — halaman lain bisa GET tanpa upload ulang
    session['site_filename'] = filename

    return jsonify({
        'success'  : True,
        'filename' : filename,
        'siteCount': len(site_index),
        'siteIndex': site_index
    })

# ================= API: GET SITE (pakai session, tanpa upload ulang) =================
@app.route('/api/get-site', methods=['GET'])
def get_site():
    """
    Halaman lain (route.html, drivetest.html, dll) tinggal GET /api/get-site
    tanpa perlu upload ulang. Data diambil dari file yang sudah disimpan.
    """
    filename = session.get('site_filename')

    if not filename:
        return jsonify({'error': 'Belum ada file site yang di-upload', 'has_site': False}), 404

    filepath = os.path.join(app.config['SHAPE_UPLOAD_FOLDER'], filename)

    if not os.path.exists(filepath):
        session.pop('site_filename', None)
        return jsonify({'error': 'File tidak ditemukan, silakan upload ulang', 'has_site': False}), 404

    try:
        site_index = parse_xlsx(filepath)
    except Exception as e:
        return jsonify({'error': f'Gagal membaca file: {str(e)}'}), 500

    return jsonify({
        'success'   : True,
        'has_site'  : True,
        'filename'  : filename,
        'siteCount' : len(site_index),
        'siteIndex' : site_index
    })

# ================= PARSE XLSX =================
# Fix v2:
#   - Blok sectors/pciList masuk ke DALAM loop for row (indentasi benar)
#   - pciList diinisialisasi bersamaan dengan sectors saat site pertama dibuat
#   - Hapus blok 'if pciList not in...' yang salah posisi
# =================================================
def parse_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    raw_header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = []
    for cell in raw_header_row:
        raw = cell.value
        cleaned = str(raw).strip() if raw is not None else ''
        headers.append(cleaned)

    clutter_candidates = ['Clutter', 'CLUTTER', 'clutter', 'CLUTTER_TYPE']
    clutter_col_found = next((h for h in headers if h in clutter_candidates), None)
    print(f"[parse_xlsx] Headers: {headers}")
    print(f"[parse_xlsx] Kolom clutter terdeteksi: '{clutter_col_found}'")

    # Deteksi kolom PCI
    pci_col_found = next((h for h in headers if h.upper() in ['PCI', 'NR_PCI', 'LTE_PCI', 'PCID']), None)
    print(f"[parse_xlsx] Kolom PCI terdeteksi: '{pci_col_found}'")

    def get_col(row_dict, *candidates):
        for key in candidates:
            if key in row_dict and row_dict[key] is not None:
                return row_dict[key]
        return None

    CLUTTER_MAP = {
        'dense urban' : {'scenario': 'umi', 'condition': 'nlos'},
        'metropolitan': {'scenario': 'umi', 'condition': 'nlos'},
        'urban'       : {'scenario': 'uma', 'condition': 'nlos'},
        'sub urban'   : {'scenario': 'uma', 'condition': 'los_nlos'},
        'suburban'    : {'scenario': 'uma', 'condition': 'los_nlos'},
        'rural'       : {'scenario': 'rma', 'condition': 'los'},
        '#n/a'        : {'scenario': 'uma', 'condition': 'nlos'},
    }

    def resolve_clutter(raw_clutter):
        if raw_clutter is None:
            return {'scenario': 'uma', 'condition': 'nlos'}
        key = ''.join(c for c in str(raw_clutter).strip().lower() if c.isprintable())
        if key in CLUTTER_MAP:
            return CLUTTER_MAP[key]
        for k in ['dense urban', 'metropolitan', 'sub urban', 'suburban', 'urban', 'rural']:
            if k in key:
                return CLUTTER_MAP[k]
        return {'scenario': 'uma', 'condition': 'nlos'}

    site_index = {}
    unmatched_clutters = set()

    # ── LOOP UTAMA — semua logika HARUS di dalam loop ini ──────────────────
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        raw_id = get_col(row_dict, 'SITE_ID', 'Site_ID', 'SITE ID', 'site_id')
        if not raw_id:
            continue

        site_id = str(raw_id).strip()

        lat_raw     = get_col(row_dict, 'LAT', 'Lat', 'LATITUDE', 'latitude')
        lng_raw     = get_col(row_dict, 'LONG', 'Long', 'LONGITUDE', 'LON', 'longitude')
        az_raw      = get_col(row_dict, 'Azimuth', 'AZIMUTH', 'azimuth')
        h_raw       = get_col(row_dict, 'Height', 'HEIGHT', 'height')
        clutter_raw = get_col(row_dict, 'CLUTTER', 'Clutter', 'clutter', 'CLUTTER_TYPE')
        # Baca PCI — coba semua nama kolom yang mungkin
        pci_raw     = get_col(row_dict, 'PCI', 'pci', 'NR_PCI', 'LTE_PCI', 'PCID')

        try:
            lat = float(lat_raw)
            lng = float(lng_raw)
        except (TypeError, ValueError):
            continue

        if not (math.isfinite(lat) and math.isfinite(lng)):
            continue

        try:
            height = float(h_raw) if h_raw is not None else 30.0
            if not math.isfinite(height):
                height = 30.0
        except (TypeError, ValueError):
            height = 30.0

        # ── Inisialisasi site baru ──────────────────────────────────────────
        if site_id not in site_index:
            clutter_info = resolve_clutter(clutter_raw)
            clutter_str  = str(clutter_raw).strip() if clutter_raw is not None else 'N/A'

            if clutter_raw is not None:
                key = str(clutter_raw).strip().lower()
                known = ['dense urban','metropolitan','sub urban','suburban','urban','rural','#n/a']
                if not any(k in key for k in known):
                    unmatched_clutters.add(clutter_str)

            site_index[site_id] = {
                'lat'      : lat,
                'lng'      : lng,
                'height'   : height,
                'sectors'  : [],     # azimuth per sektor
                'pciList'  : [],     # PCI per sektor — sejajar dengan sectors[]
                'clutter'  : clutter_str,
                'scenario' : clutter_info['scenario'],
                'condition': clutter_info['condition'],
            }
        else:
            # Update clutter jika sebelumnya N/A
            if site_index[site_id]['clutter'] == 'N/A' and clutter_raw is not None:
                clutter_info = resolve_clutter(clutter_raw)
                site_index[site_id]['clutter']   = str(clutter_raw).strip()
                site_index[site_id]['scenario']  = clutter_info['scenario']
                site_index[site_id]['condition'] = clutter_info['condition']

        # ── Tambahkan azimuth sektor + PCI — MASIH DI DALAM LOOP ───────────
        try:
            az = float(az_raw)
            if math.isfinite(az):
                site_index[site_id]['sectors'].append(int(az))

                # Baca PCI untuk sektor ini
                try:
                    pci = int(float(pci_raw)) if pci_raw is not None and str(pci_raw).strip() != '' else None
                except (TypeError, ValueError):
                    pci = None

                site_index[site_id]['pciList'].append(pci)

        except (TypeError, ValueError):
            pass
    # ── AKHIR LOOP ──────────────────────────────────────────────────────────

    wb.close()

    # Summary
    print(f"[parse_xlsx] Selesai: {len(site_index)} sites")
    if unmatched_clutters:
        print(f"[parse_xlsx] ⚠️  Clutter tidak dikenali: {unmatched_clutters}")

    from collections import Counter
    dist_clutter = Counter(d['clutter'] for d in site_index.values())
    print(f"[parse_xlsx] Distribusi clutter: {dict(dist_clutter)}")

    # Debug PCI: hitung berapa site yang punya PCI
    sites_with_pci = sum(1 for s in site_index.values() if any(p is not None for p in s.get('pciList', [])))
    total_pci      = sum(sum(1 for p in s.get('pciList', []) if p is not None) for s in site_index.values())
    print(f"[parse_xlsx] PCI: {total_pci} entri dari {sites_with_pci} site")

    # Debug sample: tampilkan 2 site pertama untuk verifikasi
    sample_ids = list(site_index.keys())[:2]
    for sid in sample_ids:
        s = site_index[sid]
        print(f"[parse_xlsx] Sample {sid}: sectors={s['sectors']}, pciList={s['pciList']}, clutter={s['clutter']}")

    return site_index

# ================= RUN =================
if __name__ == "__main__":
    app.run(debug=True)